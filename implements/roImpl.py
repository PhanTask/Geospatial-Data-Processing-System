# -*- coding:utf-8 -*-
"""
roImpl.py
~~~~~~~~~

质检报告导出功能实现。

:copyright: (c) 2018 by Jinmeng Rao.
"""

from PySide2 import QtCore
# import xlwt
from openpyxl import Workbook

class roImpl(QtCore.QThread):
    """
    导出质检报告。
    """
    processSignal = QtCore.Signal(int) # 进度signal
    isfinishedSignal = QtCore.Signal(bool)  # 是否完成

    def __init__(self, TableWidget, reportOutputPath):
        """
        初始化roImpl类。

        :param domTableWidget: 主界面的domTableWidget
        """
        super(roImpl, self).__init__()
        self.TableWidget = TableWidget
        self.reportOutputPath = reportOutputPath
        #self.wbk = xlwt.Workbook()
        self.wbk = Workbook()

    def emitIsFinished( self, isSuccessful):
        """
        发送是否完成的信号。

        :param isSuccessful: 是否成功完成，为boolean值
        :return: None
        """
        self.isfinishedSignal.emit( isSuccessful )

    def emitProcessValue(self,pValue):
        """
        发送进度条值的信号。

        :param pValue: 进度条值，为int值
        :return:
        """
        self.processSignal.emit( pValue )

    def run(self, *args, **kwargs):
        """
        在本线程中执行操作。

        :return: None
        """
        # sheet = self.wbk.add_sheet("sheet", cell_overwrite_ok=True)
        sheet = self.wbk.active
        try:
            colCount = self.TableWidget.columnCount()
            rowCount = self.TableWidget.rowCount()
            [sheet.cell(1,j+1,[self.TableWidget.horizontalHeaderItem(i).text().__str__() for i in range(colCount)][j]) for j in range(colCount)]
            # [sheet.write(0, j, [self.TableWidget.horizontalHeaderItem(i).text().__str__() for i in range(colCount)][j]) for j in range(colCount)]
            for currentRow in range(rowCount):
                for currentColumn in range(colCount):
                    text = self.TableWidget.item(currentRow, currentColumn).text()
                    sheet.cell(currentRow+2, currentColumn+1, text.__str__())
                    # sheet.write(currentRow+1, currentColumn, text.__str__())
                self.emitProcessValue(int(currentRow*100.0/rowCount))
            self.wbk.save(self.reportOutputPath)
        except Exception as e:
            print('[roImpl->run ERROR] ' + e.__str__())
            self.emitIsFinished(False)
            return
        self.emitIsFinished(True)