#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/6/4 14:26
# @Author  : Dutian
# @Site    : 
# @File    : repeatCheckImp.py
# @Software: PyCharm
# @license : Copyright(C), Dutian
# @Contact : free.du@qq.com
import os
from PySide2 import QtCore
import multiprocessing as mp
from osgeo import ogr
from logger.logRcd import logRcd
from fuzzywuzzy import fuzz
import time
#
# redefine process pool via inheritance
# import multiprocess.context as context
# class NoDaemonProcess(context.Process):
#     def _get_daemon(self):
#         return False
#     def _set_daemon(self, value):
#         pass
#     daemon = property(_get_daemon, _set_daemon)
#
# class NoDaemonPool(pathos.multiprocessing.Pool):
#     def Process(self, *args, **kwds):
#         return NoDaemonProcess(*args, **kwds)

class repeatCheckImp(QtCore.QThread):
    """
    重复性检查类
    """
    doneflag  = QtCore.Signal(bool) # 是否成功完成
    stdTableItemDataSignal = QtCore.Signal(object) # std表格项signal
    dostop = False # 是否强制终止
    processValue = 0.0 # 进度
    indexPathList = []
    stdFileNum = 0
    def __init__(self,gdbDirInputPath,gdbDirOutputPath,matchDegree):
        """
        初始化
        :param gdbDirInputPath: 输入路径
        :param gdbDirOutputPath: 输出路径
        :param matchDegree: 匹配度
        """
        super(repeatCheckImp, self).__init__()
        self.gdbDirInputPath = gdbDirInputPath
        self.gdbDirOutputPath = gdbDirOutputPath

        self.gdbInputDirPathNoExtList = []
        self.gdbOutputDirPathNoExtList = []

        self.matchDegree = matchDegree

        self.stdGdbCount=0  #gdb输入图层数量
        self.gdbCompleted=0 #gdb已处理数量
        self.layerNum = 0  # layer数量
        self.layerCompleted = 0#layer已处理数量
        self.logRcd = logRcd('时空专题大数据名称重复性检查日志')
        # self.gdbMatchDicDic=gdbMatchDicDic # 指定匹配图层，按照默认图层名来计算

    def emitFinish(self, isSuccessful):
        self.doneflag.emit(isSuccessful)

    def emitStdTableItemData(self, StdTableItemData):
        self.stdTableItemDataSignal.emit(StdTableItemData)

    def run(self, *args, **kwargs):
        """
        算法执行函数，计算文件数量，连接多进程函数
        :param args:
        :param kwargs:
        :return:
        """
        try:
            self.stdGdbCount=self.getAndCopyInputPathList()
            self.startProcess()

        except Exception as e:
            print('ERROR ' + e.__str__())


    def getAndCopyInputPathList(self):
        """
        获取gdb文件列表,返回输入文件数量列表
        :param gdbDirInputPath:
        :return: gdbOutputDirPathNoExtList
        """

        # 列表推导式  os.path.join()：  将多个路径组合后返回
        for dir in os.listdir(self.gdbDirInputPath + '/'):
            if dir[-3:] == 'gdb':
                # 复制为新的处理文件
                copyFilePath = self.gdbDirOutputPath + '/' + dir.replace('.gdb', '.xlsx')
                dir = self.gdbDirInputPath + '/' + dir
                self.gdbInputDirPathNoExtList.append(dir)
                self.gdbOutputDirPathNoExtList.append(copyFilePath)
        return self.gdbInputDirPathNoExtList.__len__()
        #return allLayerCount  # 返回图层总数


    def startProcess(self):
        """
        开启多进程函数
        :return:
        """
        self.logRcd.emitInfo('[CONFIG] 任务项：时空专题大数据名称重复性')
        self.logRcd.emitInfo('[CONFIG] 数据输入目录：' + self.gdbDirInputPath)
        self.logRcd.emitInfo('[CONFIG] 数据输出目录：' + self.gdbDirOutputPath)
        self.logRcd.emitInfo('[CONFIG] 字段模糊匹配值：' + self.matchDegree.__str__())
        self.logRcd.emitInfo('[CONFIG] 检查模式：并行（CPU核心数 '+ mp.cpu_count().__str__() +  '）')
        self.logRcd.emitInfo('[INFO] 进程池初始化...')
        processNum = mp.cpu_count() if mp.cpu_count() <= 8 else 8
        domPool = mp.Pool(processNum)
        self.logRcd.emitInfo('[INFO] 初始化成功，开始检查数据...\n')
        #driver = ogr.GetDriverByName('FileGDB')

        for gdbInputDir,gdbOutDir in zip(self.gdbInputDirPathNoExtList,self.gdbOutputDirPathNoExtList):
            driver = ogr.GetDriverByName('OpenFileGDB')
            gdbOutFile = driver.Open(gdbInputDir, 0)
            if gdbOutFile == None: return
            oLayerDic={'name': 'name', 'cityName': 'szs', 'contyName': 'szqx', 'townName': 'szz'}
            # totalDataDict = {}
            for oLayer in gdbOutFile:
                layerDict = {}
                layerName=oLayer.GetName()
                #totalDataDict[oLayerName] = {}
                for oFeature in oLayer:
                    FID = oFeature.GetFID().__str__()
                    layerDict[FID] = {}
                    layerDict[FID]['name'] = oFeature.GetField(oLayerDic['name'])
                    layerDict[FID]['cityName'] = oFeature.GetField(oLayerDic['cityName'])
                    layerDict[FID]['contyName'] = oFeature.GetField(oLayerDic['contyName'])
                    layerDict[FID]['townName'] = oFeature.GetField(oLayerDic['townName'])
                self.layerNum += 1
                domPool.apply_async(self.mpProcess, (layerDict,layerName,self.matchDegree,gdbInputDir,gdbOutDir), callback=self.mpProcessCallBack)
        domPool.close()
        domPool.join()
        self.logRcd.emitProcessValue(100)
        self.logRcd.emitInfo('[INFO] 检查完成')
        self.logRcd.logFileClose()

    @staticmethod
    def mpProcess(layerDict,layerName,matchDegree,gdbInputDir,gdbOutDir):
        """
        进程执行函数
        :param _:
        :return:
        """
        try:
            print('mpProcess entered')
            subItemDataList = []
            allFetFieldDict = {}
            for FID in layerDict:
                name = layerDict[FID]['name']
                cityName = layerDict[FID]['cityName']
                contyName = layerDict[FID]['contyName']
                townName = layerDict[FID]['townName']
                fuzzyDict = {}
                for itemFID in allFetFieldDict:
                    itemName = allFetFieldDict[itemFID]['name']
                    itemCityName = allFetFieldDict[itemFID]['cityName']
                    itemContyName = allFetFieldDict[itemFID]['contyName']
                    itemTownName = allFetFieldDict[itemFID]['townName']
                    fuzzNum = fuzz.ratio(name, itemName)
                    if fuzzNum >= matchDegree \
                            and cityName == itemCityName \
                            and contyName == itemContyName \
                            and townName == itemTownName:
                        allFetFieldDict[itemFID]['fuzzy'][FID] = fuzzNum
                        fuzzyDict[itemFID] = fuzzNum
                allFetFieldDict[FID] = {}
                allFetFieldDict[FID]['name'] = name
                allFetFieldDict[FID]['layerName'] = layerName
                allFetFieldDict[FID]['cityName'] = cityName
                allFetFieldDict[FID]['contyName'] = contyName
                allFetFieldDict[FID]['townName'] = townName
                allFetFieldDict[FID]['fuzzy'] = fuzzyDict
            # print(allFetFieldDict)
            # 准备写入
            infoString = ''
            infoString += '\n' + layerName + '\n'
            fuzzyResult_Total = ''
            print('before createsheet')
            from openpyxl import Workbook
            wbk = Workbook()
            sheet = wbk.active
            sheet.title = layerName
            print('after createsheet')
            excelRowCount = 1
            sheet.cell(row = excelRowCount,column = 1,value="ObjectID")
            sheet.cell(row = excelRowCount,column = 2,value="要素名称")
            sheet.cell(row = excelRowCount,column = 3,value="所属图层")
            sheet.cell(row = excelRowCount,column = 4,value="重复项ObjectID及相似度")
            print('after cell')
            for itemFID in allFetFieldDict:
                if allFetFieldDict[itemFID]['fuzzy'].__len__() == 0:
                    continue
                excelRowCount += 1
                itemName = allFetFieldDict[itemFID]['name']
                fuzzyResult = allFetFieldDict[itemFID]['fuzzy'].__str__()
                sheet.cell(row=excelRowCount, column=1, value=itemFID)
                sheet.cell(row=excelRowCount, column=2, value=itemName)
                sheet.cell(row=excelRowCount, column=3, value=layerName)
                sheet.cell(row=excelRowCount, column=4, value=fuzzyResult)
                infoString += '[INFO] ' + itemName + ' : ' + fuzzyResult + '\n'
                if fuzzyResult_Total.__len__() < 100:
                    fuzzyResult_Total += fuzzyResult

            if fuzzyResult_Total.__len__() >= 100:
                fuzzyResult_Total += '......'
            if fuzzyResult_Total.__len__() == 0:
                fuzzyResult_Total = '无'
            # print(fuzzyResult_Total)
            stdReTableItemData = (layerName,
                                  allFetFieldDict.__len__().__str__(),
                                  (excelRowCount - 1).__str__(),
                                  fuzzyResult_Total,
                                  gdbInputDir,
                                  gdbOutDir
                                  )
            subItemDataList.append(stdReTableItemData)
            wbk.save(gdbOutDir.replace('.xlsx', '_' + layerName + '.xlsx'))
            return [infoString,subItemDataList,gdbInputDir,layerName]
        except Exception as e:
            print(e.__str__())
        #
        #
        #     subPool = mp.Pool(4)
        #     for layerName in self.totalDataDict:
        #         subPool.apply_async(self.subProcess,(self.totalDataDict[layerName], layerName, self.matchDegree, self.gdbInputDir, self.gdbOutDir),self.subProcessCallback)
        #         #subPool.apply(self.subProcess, (self.totalDataDict[layerName], layerName, self.matchDegree, self.gdbInputDir, self.gdbOutDir))
        #         print('subPool applied')
        #     subPool.close()
        #     subPool.join()
        #     # self.wbk.save(self.gdbOutDir)
        #     return [self.infoString,self.itemDataList,self.gdbInputDir]
        # except Exception as e:
        #     print(e.__str__())

    def mpProcessCallBack(self, res):
        """
        回调函数
        :param res: 列表,res[0]为infoString，res[1]为表格项列表, res[2]为InputDir路径, res[3]为图层名
        :return:
        """
        print('mpProcessCallBack')
        # self.gdbCompleted += 1
        # self.logRcd.emitInfo('[INFO] 正在生成质检日志与质检报告，请稍候...')

        self.layerCompleted += 1

        self.logRcd.emitProcessValue(int(100.0 * (self.layerCompleted / self.layerNum)))

        self.logRcd.emitInfo('[INFO] ' + res[2])
        self.logRcd.emitInfo('(' + self.layerCompleted.__str__() + '/' + self.layerNum.__str__() + ') 图层名：' + res[3])
        self.logRcd.emitInfo(res[0])
        self.emitStdTableItemData(res[1])

    # @staticmethod
    # def mpProcess(oLayerDic,totalDataDict,gdbInputDir,gdbOutDir,matchDegree):
    #     """
    #     单进程执行函数
    #     :param gdbOutFile:
    #     :param matchDegree:
    #     :return:
    #     """
    #     subPool = mp.Pool(mp.cpu_count())
    #     # try:
    #         itemDataList = []
    #         infoString = ''
    #
    #         excelRowCount = 1
    #
    #         for layerName in totalDataDict:
    #             allFetFieldDict = {}
    #             subPool.apply_async(subProcess,(totalDataDict[layerName],layerName,matchDegree),subProcessCallback)
    #
    #         subPool.close()
    #         subPool.join()
    #             #准备写入
    #             infoString += '\n' + layerName + '\n'
    #             fuzzyResult_Total = ''
    #             for itemFID in allFetFieldDict:
    #                 if allFetFieldDict[itemFID]['fuzzy'].__len__() == 0:
    #                     continue
    #                 excelRowCount += 1
    #                 itemName = allFetFieldDict[itemFID]['name']
    #                 fuzzyResult = allFetFieldDict[itemFID]['fuzzy'].__str__()
    #                 sheet.cell(row=excelRowCount, column=1, value=itemFID)
    #                 sheet.cell(row=excelRowCount, column=2, value=itemName)
    #                 sheet.cell(row=excelRowCount, column=3, value=layerName)
    #                 sheet.cell(row=excelRowCount, column=4, value=fuzzyResult)
    #                 infoString += '[INFO] ' + itemName + ' : ' + fuzzyResult + '\n'
    #                 fuzzyResult_Total += fuzzyResult
    #             print(fuzzyResult_Total)
    #             stdReTableItemData = (layerName,
    #                                   oLayer.GetFeatureCount().__str__(),
    #                                   (excelRowCount-1).__str__(),
    #                                   fuzzyResult_Total,
    #                                   gdbInputDir,
    #                                   gdbOutDir
    #                                   )
    #             itemDataList.append(stdReTableItemData)
                # allFeatRecordDicListDic={}
                #
                # # 遍历list找到重复值
                # for i in range(len(allFetFieldList)):
                #     oFeatName = allFetFieldList[i]['name']
                #     oFeatID=allFetFieldList[i]['FID']
                #
                #     initFid = allFetFieldList[i]['FID']
                #     for j in range(i + 1, len(allFetFieldList)):
                #         fuzzNum = fuzz.ratio(allFetFieldList[j]['name'], oFeatName)
                #         # fuzzNum = fuzz._token_sort(allFetFieldList[j]['name'], oFeatName,
                #         #                            partial=False,
                #         #                            force_ascii=True,
                #         #                            full_process=True)
                #         if fuzzNum >= matchDegree \
                #                 and allFetFieldList[j]['cityName'] == allFetFieldList[i]['cityName'] \
                #                 and allFetFieldList[j]['contyName'] == allFetFieldList[i]['contyName'] \
                #                 and allFetFieldList[j]['townName'] == allFetFieldList[i]['townName']:
                #             compareJFid=allFetFieldList[j]['FID']
                #
                #             allFeatRecordDicList = []
                #             allFeatRecordDicList1 = []
                #             oFeatRecordDic = {}
                #             oFeatRecordDic1 = {}  # 反向存储
                #
                #             oFeatRecordDic[compareJFid]=fuzzNum
                #             oFeatRecordDic1[initFid]=fuzzNum
                #             allFeatRecordDicList.append(oFeatRecordDic)
                #             allFeatRecordDicList1.append(oFeatRecordDic1)
                #
                #             if oFeatID in allFeatRecordDicListDic.keys():
                #                 allFeatRecordDicListDic[oFeatID].extend(allFeatRecordDicList)
                #             else:
                #                 allFeatRecordDicListDic[oFeatID]=allFeatRecordDicList
                #
                #             if compareJFid in allFeatRecordDicListDic.keys():
                #                 allFeatRecordDicListDic[compareJFid].extend(allFeatRecordDicList1)
                #             else:
                #                 allFeatRecordDicListDic[compareJFid]=allFeatRecordDicList1
                #
                #             #oFeature.SetField(u"correct", sortResultStr)
                #             # print(oFeatName +'-' \
                #             #     +allFetFieldList[j]['name'])

                # #准备写入字段
                # infoString += '\n' + oLayerName + '\n'
                # # self.logRcd.emitInfo('\n' + oLayer.GetName())
                # oLayer.ResetReading()
                # itemCount = 1
                # for oFeatureWrite in oLayer:
                #     featFID = oFeatureWrite.GetFID()
                #     featName = oFeatureWrite.GetField('name')
                #     if featFID in allFeatRecordDicListDic.keys():
                #         excelRowCount += 1
                #         oFeatDicList= allFeatRecordDicListDic[featFID]
                #         # oFeatureWrite.SetField(u'repeatCheckField',oFeatDicList.__str__())
                #         # oLayer.SetFeature(oFeatureWrite)
                #         sheet.cell(row=excelRowCount, column=1, value=featFID.__str__())
                #         sheet.cell(row=excelRowCount, column=2, value=featName)
                #         sheet.cell(row=excelRowCount, column=3, value=oLayerName)
                #         sheet.cell(row=excelRowCount, column=4, value=oFeatDicList.__str__())
                #         # print(oFeatureWrite.GetField('name'),oFeatDicList)
                #         # self.logRcd.emitInfo('[INFO] ' + oFeatureWrite.GetField('name') + ' : ' + oFeatDicList.__str__())
                #         infoString += '[INFO] ' + featName + ' : ' + oFeatDicList.__str__() + '\n'
                #         print(oFeatDicList.__str__())
                #         itemCount += 1
                #     #oLayer.SetFeature(oFeatureWrite.GetFID())
                # # self.logRcd.emitProcessValue(int(100.0 * (((self.gdbCompleted - 1) / self.stdGdbCount) + layerCount / layerNum / self.stdGdbCount)))
                # #质检表格内容  格式为：图层名称，要素总个数，重复性名称要素个数，重名要素FID字符串，输入路径，输出路径
                # repeatFeatFIDStr= ','.join( str(key) for key in allFeatRecordDicListDic.keys())
                # print(repeatFeatFIDStr)
                # stdReTableItemData=(oLayerName,
                #                     oLayer.GetFeatureCount().__str__(),
                #                     len(allFeatRecordDicListDic).__str__(),
                #                     repeatFeatFIDStr,
                #                     gdbInputDir,
                #                     gdbOutDir
                #                     )
                # itemDataList.append(stdReTableItemData)
                # # self.emitStdTableItemData(stdReTableItemData)
                #
                # #self.logRcd.emitInfo(allFeatRecordDicListDic.__str__())
                # # 表格信号打包发送
                # # stdReTableItemDataList.append(stdReTableItemData)
                # # 进程信号为单图层所有要素重复统计信息
                # # signalProcessList.append(allFeatRecordDicListDic)
                # layerCount += 1
        #     wbk.save(gdbOutDir)
        #     print('saved')
        #     return (infoString,itemDataList,gdbInputDir)
        # except Exception as e:
        #     print(e.__str__())
# class subProcessWorker():
#
#     def __init__(self,oLayerDic,totalDataDict,gdbInputDir,gdbOutDir,matchDegree):
#         self.oLayerDic = oLayerDic
#         self.totalDataDict = totalDataDict
#         self.gdbInputDir = gdbInputDir
#         self.gdbOutDir = gdbOutDir
#         self.matchDegree = matchDegree
#         self.itemDataList = []
#         self.infoString = ''
#
#     def mpProcess(self):
#         """
#         进程执行函数
#         :param _:_
#         :return:
#         """
#         try:
#             print('mpProcess entered')
#             subPool = mp.Pool(4)
#             for layerName in self.totalDataDict:
#                 subPool.apply_async(self.subProcess,(self.totalDataDict[layerName], layerName, self.matchDegree, self.gdbInputDir, self.gdbOutDir),self.subProcessCallback)
#                 #subPool.apply(self.subProcess, (self.totalDataDict[layerName], layerName, self.matchDegree, self.gdbInputDir, self.gdbOutDir))
#                 print('subPool applied')
#             subPool.close()
#             subPool.join()
#             # self.wbk.save(self.gdbOutDir)
#             return [self.infoString,self.itemDataList,self.gdbInputDir]
#         except Exception as e:
#             print(e.__str__())
#
#     @staticmethod
#     def subProcess(layerDict,layerName,matchDegree,gdbInputDir,gdbOutDir):
#         print('subProcess')
#         try:
#             subItemDataList = []
#             allFetFieldDict = {}
#             for FID in layerDict:
#                 name = layerDict[FID]['name']
#                 cityName = layerDict[FID]['cityName']
#                 contyName = layerDict[FID]['contyName']
#                 townName = layerDict[FID]['townName']
#                 fuzzyDict = {}
#                 for itemFID in allFetFieldDict:
#                     itemName = allFetFieldDict[itemFID]['name']
#                     itemCityName = allFetFieldDict[itemFID]['cityName']
#                     itemContyName = allFetFieldDict[itemFID]['contyName']
#                     itemTownName = allFetFieldDict[itemFID]['townName']
#                     fuzzNum = fuzz.ratio(name, itemName)
#                     if fuzzNum >= matchDegree \
#                             and cityName == itemCityName \
#                             and contyName == itemContyName \
#                             and townName == itemTownName:
#                         allFetFieldDict[itemFID]['fuzzy'][FID] = fuzzNum
#                         fuzzyDict[itemFID] = fuzzNum
#                 allFetFieldDict[FID] = {}
#                 allFetFieldDict[FID]['name'] = name
#                 allFetFieldDict[FID]['layerName'] = layerName
#                 allFetFieldDict[FID]['cityName'] = cityName
#                 allFetFieldDict[FID]['contyName'] = contyName
#                 allFetFieldDict[FID]['townName'] = townName
#                 allFetFieldDict[FID]['fuzzy'] = fuzzyDict
#             print(allFetFieldDict)
#             #return [allFetFieldDict,layerName]
#             res = [allFetFieldDict,layerName]
#             print('subProcessCallback')
#             # 准备写入
#             infoString = ''
#             infoString += '\n' + res[1] + '\n'
#             fuzzyResult_Total = ''
#             print('before createsheet')
#             from openpyxl import Workbook
#             wbk = Workbook()
#             sheet = wbk.active
#             sheet.title = res[1]
#             print('after createsheet')
#             excelRowCount = 1
#             sheet.cell(row = excelRowCount,column = 1,value="ObjectID")
#             sheet.cell(row = excelRowCount,column = 2,value="要素名称")
#             sheet.cell(row = excelRowCount,column = 3,value="所属图层")
#             sheet.cell(row = excelRowCount,column = 4,value="重复项ObjectID及相似度")
#             print('after cell')
#
#             for itemFID in res[0]:
#                 if res[0][itemFID]['fuzzy'].__len__() == 0:
#                     continue
#                 excelRowCount += 1
#                 itemName = res[0][itemFID]['name']
#                 fuzzyResult = res[0][itemFID]['fuzzy'].__str__()
#                 sheet.cell(row=excelRowCount, column=1, value=itemFID)
#                 sheet.cell(row=excelRowCount, column=2, value=itemName)
#                 sheet.cell(row=excelRowCount, column=3, value=res[1])
#                 sheet.cell(row=excelRowCount, column=4, value=fuzzyResult)
#                 infoString += '[INFO] ' + itemName + ' : ' + fuzzyResult + '\n'
#                 fuzzyResult_Total += fuzzyResult
#             print(fuzzyResult_Total)
#             stdReTableItemData = (res[1],
#                                   res[0].__len__().__str__(),
#                                   (excelRowCount - 1).__str__(),
#                                   fuzzyResult_Total,
#                                   gdbInputDir,
#                                   gdbOutDir
#                                   )
#             subItemDataList.append(stdReTableItemData)
#             wbk.save(gdbOutDir.replace('.xlsx', '_' + res[1] + '.xlsx'))
#             return [subItemDataList,infoString]
#         except Exception as e:
#             print(e.__str__())
#
#     def subProcessCallback(self,res):
#         print('subProcessCallback')
#         try:
#             self.infoString += res[1]
#             self.itemDataList.extend(res[0])
#             # # 准备写入
#             # self.infoString += '\n' + res[1] + '\n'
#             # fuzzyResult_Total = ''
#             # print('before createsheet')
#             # from openpyxl import Workbook
#             # wbk = Workbook()
#             # sheet = wbk.active
#             # sheet.title = res[1]
#             # print('after createsheet')
#             # excelRowCount = 1
#             # sheet.cell(row = excelRowCount,column = 1,value="ObjectID")
#             # sheet.cell(row = excelRowCount,column = 2,value="要素名称")
#             # sheet.cell(row = excelRowCount,column = 3,value="所属图层")
#             # sheet.cell(row = excelRowCount,column = 4,value="重复项ObjectID及相似度")
#             # print('after cell')
#             #
#             # for itemFID in res[0]:
#             #     if res[0][itemFID]['fuzzy'].__len__() == 0:
#             #         continue
#             #     excelRowCount += 1
#             #     itemName = res[0][itemFID]['name']
#             #     fuzzyResult = res[0][itemFID]['fuzzy'].__str__()
#             #     sheet.cell(row=excelRowCount, column=1, value=itemFID)
#             #     sheet.cell(row=excelRowCount, column=2, value=itemName)
#             #     sheet.cell(row=excelRowCount, column=3, value=res[1])
#             #     sheet.cell(row=excelRowCount, column=4, value=fuzzyResult)
#             #     self.infoString += '[INFO] ' + itemName + ' : ' + fuzzyResult + '\n'
#             #     fuzzyResult_Total += fuzzyResult
#             # print(fuzzyResult_Total)
#             # stdReTableItemData = (res[1],
#             #                       res[0].__len__().__str__(),
#             #                       (excelRowCount - 1).__str__(),
#             #                       fuzzyResult_Total,
#             #                       self.gdbInputDir,
#             #                       self.gdbOutDir
#             #                       )
#             # self.itemDataList.append(stdReTableItemData)
#             # wbk.save(self.gdbOutDir.replace('.xlsx', '_' + res[1] + '.xlsx'))
#         except Exception as e:
#             print(e.__str__())