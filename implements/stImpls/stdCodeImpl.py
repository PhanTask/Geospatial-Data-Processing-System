# -*- coding: utf-8 -*-
#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2018/6/5 10:46
# @Author  : Dutian
# @Site    :
# @File    : codeConsistCheckImp.py
# @Software: PyCharm
# @license : Copyright(C), Dutian
# @Contact : free.du@qq.com

import os
import multiprocessing as mp
from PySide2 import QtCore
from logger.logRcd import logRcd
import ogr
from openpyxl import Workbook

class stdCodeImpl(QtCore.QThread):
    """
    处理std任务
    """
    doneflag  = QtCore.Signal(bool) # 是否成功完成
    stdTableItemDataSignal = QtCore.Signal(object) # std表格项signal
    dostop = False # 是否强制终止
    processValue = 0.0 # 进度
    stdFileNum = 0
    def __init__(self, indexDirPath, outputDirPath = None):
        """
        初始化stdCodeImpl类

        :param indexDirPath: 索引文件夹路径，为str
        """
        super(stdCodeImpl, self).__init__()
        self.indexDirPath = indexDirPath
        self.outputDirPath = outputDirPath
        self.logRcd = logRcd('时空专题大数据编码一致性检查日志')
        self.indexPathList = []
        self.outputPathList = []
        self.indexFileNum = 0
        self.stdCount = 0

    def emitFinish( self, isSuccessful ):
        self.doneflag.emit( isSuccessful )

    def emitStdTableItemData(self, StdTableItemData):
        self.stdTableItemDataSignal.emit(StdTableItemData)

    def run(self, *args, **kwargs):
        """
        在本线程中执行算法。

        :return: None
        """
        try:
            self.indexFileNum = self.getIndexPathList()
            self.stdCodeChecker(self.indexFileNum)
        except Exception as e:
            print('ERROR ' + e.__str__())

    def getIndexPathList(self):
        """
        从所给的索引文件夹路径中将所有索引的路径提取出来。

        :return: 索引数据数量，为int
        """
        self.indexPathList = [ self.indexDirPath + '/' + dir for dir in os.listdir(self.indexDirPath)]
        self.outputPathList = [self.outputDirPath + '/' + dir.replace('.gdb', '.xlsx') for dir in
                               os.listdir(self.indexDirPath)]
        return self.indexPathList.__len__()

    @staticmethod
    def pystdCodeChecker(indexPath,outputPath,indexFileNum):
        """
        执行编码一致性检查操作

        :param indexPath: 原始索引表文件路径名，为str
        :param indexFileNum: 原始索引表文件个数，为int
        :return resDict: 检查结果，为dict
        """

        gdbDir = indexPath
        driver = ogr.GetDriverByName('FileGDB')
        gdbFile = driver.Open(gdbDir, 0)

        nameSet = set()
        nameAndCodeSet = set()
        featDifferCounterDict = {}  # 各要素的编码种类的数量，若为1则该要素再各个图层中出现的编码都是一致的，大于1则存在不一致的情况
        resultDict = {}

        # 读取所有要素属性信息
        for oLayer in gdbFile:
            layerName = oLayer.GetName()
            for oFeature in oLayer:
                featName = oFeature.GetField('name')
                if featName not in nameSet:
                    featCode = oFeature.GetField('uuid')
                    nameSet.add(featName)
                    nameAndCodeSet.add(featName + '/' + featCode)
                    featDifferCounterDict[featName] = 1
                    resultDict[featName] = []
                    basicItem = {'name': featName,
                                 'code': featCode,
                                 'layerName': layerName,
                                 'gdbDir': gdbDir
                                 }
                    resultDict[featName].append(basicItem)
                else:
                    featCode = oFeature.GetField('uuid')
                    featNameAndCode = featName + '/' + featCode
                    if featNameAndCode not in nameAndCodeSet:  # 相同名称要素编码不一致，通常是第一次检查到不一致的情况
                        nameAndCodeSet.add(featNameAndCode)
                        featDifferCounterDict[featName] += 1
                        unconsistItem = {'name': featName,
                                         'code': featCode,
                                         'layerName': layerName,
                                         'gdbDir': gdbDir
                                         }
                        resultDict[featName].append(unconsistItem)
                    else:  # 存在相同名称要素编码一致的情况，但可能只是与部分要素的编码一致，整体还可能不一致
                        if featDifferCounterDict[featName] > 1:  # 若要素数量大于1，说明该要素编码本身就存在多个不一致情况，仍判作不一致
                            unconsistItem = {'name': featName,
                                             'code': featCode,
                                             'layerName': layerName,
                                             'gdbDir': gdbDir
                                             }
                            resultDict[featName].append(unconsistItem)
                        else:  # 遇到了相同名称要素编码完全一致的理想情况，检查合格
                            pass

        infoString = ''
        itemCount = 1
        excelRowCount = 1
        itemNum = resultDict.__len__()

        wbk = Workbook()
        sheet = wbk.active
        sheet.title = '质检结果'
        sheet.cell(row = excelRowCount,column = 1,value="要素名称")
        sheet.cell(row = excelRowCount,column = 2,value="所属图层")
        sheet.cell(row = excelRowCount,column = 3,value="要素编码")
        sheet.cell(row = excelRowCount,column = 4,value="是否一致")

        for item in resultDict:
            # time.sleep(0.005)
            infoString+='[INFO]   ' + itemCount.__str__() + '/' + itemNum.__str__() + '  ' + item.__str__() + '\n'
            if resultDict[item].__len__() == 1:
                excelRowCount += 1
                itemCount += 1
                sheet.cell(row = excelRowCount,column = 1,value=resultDict[item][0]['name'].__str__())
                sheet.cell(row = excelRowCount,column = 2,value=resultDict[item][0]['layerName'].__str__())
                sheet.cell(row = excelRowCount,column = 3,value=resultDict[item][0]['code'].__str__())
                sheet.cell(row = excelRowCount,column = 4,value='一致')
                infoString +='[INFO] 该要素编码均一致\n'
                infoString +='[INFO] 要素编码：' + resultDict[item][0]['code'].__str__() + '\n\n'
                continue
            else:
                itemCount += 1
                infoString += '[INFO] 共出现 ' + resultDict[item].__len__().__str__() + ' 处编码不一致的要素 \n'
                conCounter = 1
                for con in resultDict[item]:
                    excelRowCount += 1
                    sheet.cell(row = excelRowCount,column = 1,value=con['name'].__str__())
                    sheet.cell(row = excelRowCount,column = 2,value=con['layerName'].__str__())
                    sheet.cell(row = excelRowCount,column = 3,value=con['code'].__str__())
                    sheet.cell(row = excelRowCount,column = 4,value='不一致')
                    infoString +='[INFO] ----第 ' + conCounter.__str__() + ' 处要素----' + '\n'
                    infoString +='[INFO] 要素所在图层名称：' + con['layerName'].__str__()+ '\n'
                    infoString +='[INFO] 要素编码：' + con['code'].__str__()+ '\n'
                    infoString +='[INFO] 要素所在文件路径：' + con['gdbDir'].__str__() + '\n\n'
                    conCounter += 1
        # print(infoString)
        wbk.save(outputPath)
        return (resultDict,indexPath,indexFileNum,infoString)

    def pystdCodeCheckerCallBack(self,res):
        """
        pystdCodeChecker的回调函数，用于记录日志和更新进度条。

        :param res: pystdCodeChecker的返回值，为(resDict,indexPath,indexFileNum,infoString)
        :return: None
        """
        self.logRcd.emitInfo('[INFO] 正在生成质检日志与质检报告，请稍候...')
        print('1')
        self.stdCount += 0.5

        self.logRcd.emitProcessValue(int(100.0 * (self.stdCount / res[2])))

        totalInfoString = '[INFO] '+'(' + self.stdCount.__str__()  + '/' + res[2].__str__() + ') ' + res[1] +'\n\n' + res[3]

        self.logRcd.emitInfo(totalInfoString)
        print('2')
        self.stdCount += 0.5
        self.logRcd.emitProcessValue(int(100.0 * (self.stdCount / res[2])))
        print('3')
        self.emitStdTableItemData(res[0])
        print('4')

    def stdCodeChecker(self,indexFileNum):
        """
        时空专题大数据编码一致性检查

        :return: None
        """
        self.logRcd.emitInfo('[CONFIG] 任务项：时空专题大数据编码一致性检查')
        self.logRcd.emitInfo('[CONFIG] 索引数据输入目录：' + self.indexDirPath)
        coreNum = mp.cpu_count()
        self.logRcd.emitInfo('[CONFIG] 检查模式：并行（CPU核心数 '+ coreNum.__str__() +  '）')
        self.logRcd.emitInfo('[INFO] 进程池初始化...')
        domPool = mp.Pool(4)
        self.logRcd.emitInfo('[INFO] 初始化成功，开始检查数据...\n')

        gdbCount = 1
        for indexPath,outputPath in zip(self.indexPathList,self.outputPathList):
            #domPool.apply(self.pystdCodeChecker,(indexPath, indexFileNum))
            domPool.apply_async(self.pystdCodeChecker,(indexPath, outputPath, indexFileNum),callback=self.pystdCodeCheckerCallBack)
            self.logRcd.emitInfo('[INFO] (' + gdbCount.__str__()  + '/' +  self.indexFileNum.__str__() + ') ' +  indexPath.__str__() + ' 文件已加载，质检结果将批量返回，请稍候...\n')
            gdbCount += 1
        domPool.close()
        domPool.join()
        self.logRcd.emitProcessValue(100)
        self.logRcd.emitInfo('[INFO] 检查完成')
        self.logRcd.logFileClose()